# -*- coding: utf-8 -*-
# Corrige la asignación de grupos en rrhh_empleados (uso único).
# Fuente de verdad: hoja "CONTRATISTAS Y EMPLEADOS DNI".
#   - El JEFE de cada cuadrilla está marcado en AMARILLO (fill FFFFFF00).
#   - Todas las filas siguientes (no amarillas) hasta el próximo amarillo son su cuadrilla.
#   - Si la celda del jefe tiene el DNI roto/vacío, se resuelve por nombre contra
#     la hoja "DNI y CELULAR CONTRATISTA".
import re, difflib
import pandas as pd
from openpyxl import load_workbook

SRC = r"D:\Descargas 2\contratistas.xlsx"
OUT = r"D:\proyectos\klasea-stock\_rrhh_fix.sql"

# DNIs que en el Excel vienen rotos/vacíos en la fila del jefe (override seguro)
OVERRIDE = {
    "fritz leonardo": "21672560",
    "loys carlos": "47071647",
    "lois carlos": "47071647",
    "ramirez j gabriel": "28046196",
    "ramirez jgabriel": "28046196",
}

def norm(s):
    s = str(s or "").replace("�", "ñ")
    s = "".join(c for c in __import__("unicodedata").normalize("NFD", s) if not __import__("unicodedata").combining(c))
    s = re.sub(r"[^a-z0-9 ]", " ", s.lower())
    return re.sub(r"\s+", " ", s).strip()

def clean_name(s):
    if not isinstance(s, str): return None
    s = s.replace("�", "ñ"); s = re.sub(r"\s+", " ", s).strip()
    s = re.sub(r"\s*\d[\d\.\-]*$", "", s).strip()
    if not s: return None
    return s.replace("Luius", "Luis").replace("Loys", "Lois").title()

def dni_of(v):
    if v is None or (isinstance(v, float) and pd.isna(v)): return None
    if isinstance(v, (int, float)): return str(int(v))
    s = str(v).strip().replace(" ", "").replace("-", "")
    if not s: return None
    if re.fullmatch(r"\d+\.0", s): s = s[:-2]
    s = s.replace(".", "")
    return str(int(s)) if (s.isdigit() and 6 <= len(s) <= 9) else None

def esc(s): return str(s).replace("'", "''")

def is_yellow(cell):
    f = cell.fill
    if f.fill_type != "solid": return False
    c = f.fgColor
    return c.type == "rgb" and str(c.rgb).upper() in {"FFFFFF00", "FFFF00"}

# --- jefes del listado de celulares (dni -> nombre, dni -> celular) ---
xl = pd.read_excel(SRC, sheet_name=None, header=None)
jefes24, cel_by_dni = {}, {}
for _, r in xl["DNI y CELULAR CONTRATISTA"].iterrows():
    nombre, dni, cel = clean_name(r[1]), dni_of(r[2]), r[4]
    if nombre and dni and nombre.upper() != "NOMBRE":
        jefes24[dni] = nombre
        if pd.notna(cel):
            try: cel_by_dni[dni] = str(int(float(cel)))
            except Exception: pass
jefes24_norm = {dni: norm(n) for dni, n in jefes24.items()}

def resolver_jefe_dni(name):
    """Match del nombre del jefe amarillo contra los 24 jefes del listado."""
    nn = norm(name)
    if nn in OVERRIDE: return OVERRIDE[nn]
    htoks = {t for t in nn.split() if len(t) >= 3}
    best, bestscore = None, 0.0
    for dni, jn in jefes24_norm.items():
        jtoks = {t for t in jn.split() if len(t) >= 3}
        overlap = sum(1 for h in htoks if any(h == j or (h[:4] == j[:4]) for j in jtoks))
        score = overlap + difflib.SequenceMatcher(None, nn, jn).ratio()
        if score > bestscore: best, bestscore = dni, score
    return best if bestscore >= 1.2 else None

# --- recorrer la hoja: amarillo = jefe, resto = cuadrilla ---
wb = load_workbook(SRC, data_only=True)
ws = wb["CONTRATISTAS Y EMPLEADOS DNI"]
contratistas = {}          # dni -> nombre
worker_to_jefe = {}        # worker_dni -> jefe_dni
resoluciones = []          # para verificación
cur = None

for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
    cell = row[0]
    name = clean_name(cell.value)
    dni = dni_of(row[1].value)
    if is_yellow(cell) and name:
        jefe_dni = dni or resolver_jefe_dni(name)
        if not jefe_dni:
            resoluciones.append((name, "SIN_DNI", "??")); cur = None; continue
        jefe_name = jefes24.get(jefe_dni, name)
        contratistas.setdefault(jefe_dni, jefe_name)
        worker_to_jefe[jefe_dni] = jefe_dni
        cur = jefe_dni
        resoluciones.append((name, jefe_dni, "cell" if dni else "match"))
    elif cur and dni:                      # cuadrilla (alcanza con el DNI)
        worker_to_jefe.setdefault(dni, cur)

# jefes del listado que no aparecieron amarillos: contratista de sí mismos
for dni, nombre in jefes24.items():
    contratistas.setdefault(dni, nombre)
    worker_to_jefe.setdefault(dni, dni)

# --- SQL ---
by_jefe = {}
for w, j in worker_to_jefe.items():
    by_jefe.setdefault(j, set()).add(w)

L = ["-- ===== FIX asignacion contratistas (jefe = fila AMARILLA del Excel) ====="]
L.append("-- 1) Alta de contratistas faltantes (por DNI)")
L.append("insert into rrhh_contratistas (nombre, dni, celular) values")
vals = [f"  ('{esc(n)}', '{d}', {('NULL' if d not in cel_by_dni else chr(39)+cel_by_dni[d]+chr(39))})"
        for d, n in sorted(contratistas.items(), key=lambda x: x[1])]
L.append(",\n".join(vals)); L.append("on conflict (dni) do nothing;"); L.append("")
L.append("-- 2) Reset: todos a CASA (sin tocar sede / ficha / activo)")
L.append("update rrhh_empleados set grupo='casa', contratista_id=null;"); L.append("")
L.append("-- 3) Asignar cada empleado (por DNI) a su contratista")
for jefe_dni in sorted(by_jefe, key=lambda d: contratistas.get(d, "")):
    inlist = ",".join(f"'{d}'" for d in sorted(by_jefe[jefe_dni]))
    L.append(f"update rrhh_empleados set grupo='contratista', "
             f"contratista_id=(select id from rrhh_contratistas where dni='{jefe_dni}') where dni in ({inlist});")

with open(OUT, "w", encoding="utf-8") as f:
    f.write("\n".join(L))

print(f"OK -> {OUT}")
print(f"contratistas={len(contratistas)}  empleados_mapeados={len(worker_to_jefe)}")
print("=== resolución de jefes amarillos ===")
for name, dni, how in resoluciones:
    print(f"  {name:<26} -> {dni:<12} ({how})  crew={len(by_jefe.get(dni, [])) - 1 if dni in by_jefe else 0}")
